# -*- coding: utf-8 -*-

#! python3

# USAGE
# python3 excel_2_csv_041018_1.py "FOLDERWITHEXCEL"
# python3 excel_2_csv_041018_1.py "../tests/testExcelSheets"

import os,sys,csv,openpyxl

try: 
    from openpyxl.cell import column_index_from_string,get_column_letter
except ImportError:
    from openpyxl.utils import column_index_from_string,get_column_letter

import logging
logging.basicConfig(level=logging.DEBUG, format=" %(asctime)s - %(levelname)s - %(message)s")
# logging.disable(logging.CRITICAL)

#####################################
# COMMAND LINE PARSING
#####################################

input_folder = sys.argv[1] # should be "FOLDERWITHEXCEL", which should be a path
logging.debug('The input folder targeted is:  %s' % (input_folder))

#####################################
# END COMMAND LINE PARSING
#####################################

#####################################
# PROCESS FILES
#####################################

def convert_sheet_to_csv_1(excelFile,sheetItem,filename):
	# create output folder
	# if the folder doesn't exist, create it
	if not os.path.exists('output_csv_f'):
		os.makedirs('output_csv_f')

	# get the sheet data from the Excel file
	sheetData = excelFile[sheetItem] # where "excelFile" is the workbook, where "sheetItem" is a string, the name of the sheet
	logging.debug('sheetData is:  ')
	logging.debug(sheetData)

	# create the CSV filename from Excel filename and sheet title
	# filenames of CSV should be `<excel filename>_<sheet title>.csv`
	# csv_filename_is = filename + "_" + sheetItem.title + ".csv"
	csv_filename_is = str(filename) + "_" + str(sheetData.title) + ".csv" # this version puts the final file in the `output_csv_f` folder
	logging.debug('csv filename to use:  %s' % (csv_filename_is))
	# create csv.writer object for this CSV file
	output_csv_is = open(os.path.join("output_csv_f/",csv_filename_is),'w',newline='')
	logging.debug('output_csv_is created')
	outputWriter = csv.writer(output_csv_is)
	logging.debug('outputWriter ready')
	
	# loop through every row in the sheet data
	for rowNum in range(1,sheetData.max_row + 1):
		rowData = [] # append each cell to this list
		# loop through each cell in the row
		for colNum in range(1,sheetData.max_column + 1):
			# get the cell coordinate, this is a string
			# convert column number into a letter
			cell_coordinate = get_column_letter(colNum) + str(rowNum)
			logging.debug('cell_coordinate is:  %s' % (cell_coordinate) )
			# append each cell's data to rowData
			rowData.append(sheetData[cell_coordinate].value)
			logging.debug('rowData to write is:  ')
			logging.debug(rowData)
		# write the rowData list to the CSV file
		outputWriter.writerow(rowData)
		logging.debug('rowData written')

	# close `output_csv_is` file
	output_csv_is.close()
	logging.debug('output_csv_is closed')

for file in os.listdir(input_folder):
	# check that file is an Excel file
	if file.endswith('xlsx'):
		# create correct file path to open
		filename_path = os.path.join(input_folder,file)
		# read the file
		logging.debug('Opening file:  %s' % (file))
		excelFile = openpyxl.load_workbook(filename_path) # this has to open the `filename_path` not `file`
		logging.debug('File opened.')
		
		# create list of sheet names to target
		sheetList = excelFile.sheetnames # creates list of all sheets in the workbook
		logging.debug('List of target sheets:')
		logging.debug(sheetList)

		# loop through each sheet in the workbook
		for sheetItem in sheetList:
			convert_sheet_to_csv_1(excelFile,sheetItem,file)
	else:
		pass

#####################################
# END PROCESS FILES
#####################################

